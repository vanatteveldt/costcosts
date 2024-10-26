import csv
import datetime
import re
from openpyxl import load_workbook

vuref = "723020"
date = datetime.datetime.now().strftime("%d-%m-%Y")
toelichting = "Declaration from eCOST export"
declaratieregel = "Declaration from eCOST export"

print(">>> Reading export.csv")

rows = []
for line in csv.DictReader(open("export.csv"), delimiter=";"):
    # Naam
    name = line["\ufeffBeneficiary"]
    title, *names, ln = name.split(" ")
    initials = "".join(name[0] + "." for name in names)
    row = [initials, ln]

    # Adres
    adres = line["Account Holder Street"].strip()
    stad = line["Account Holder City"].strip()
    post = line["Account Holder Postal Code"].strip()
    tel = line["Participant Phone"].strip().replace(" ", "")
    land = line["Account Holder Country"].strip()
    email = line["Participant E-mail"].strip()
    if m := re.match(r"(\d+)\s+([A-Za-z].*)", adres):
        hn, straat = m.groups()
    elif m := re.match(r"([A-Za-z].*)\s+(\d+)", adres):
        straat, hn = m.groups()
    else:
        print("WARNING: Could not parse address ", repr(adres))
        hn, straat = None, adres
    if not email:
        print("WARNING: no email for ", name)
    row += [straat, hn, post, stad, tel, land, email, None, None]

    # Financial
    bankland = line["Bank Country"]
    rekeninghouder = line["Account Holder"]
    bic = line["SWIFT"]
    iban = line["IBAN"]
    row += [bankland, rekeninghouder, bic, iban, None, None, None]

    # Reference
    ref = line["Reference"]
    row += [vuref, ref, toelichting, date]

    # Money
    valuta, amount = line["Amount"].split(" ", 1)
    row += [valuta, declaratieregel, amount]
    rows.append(row)


print("\n>>> Reading template.xlsx")

# Open an xlsx
wb = load_workbook(filename="template.xlsx")
ws = wb[wb.sheetnames[0]]

for i, row in enumerate(rows):
    for j, value in enumerate(row):
        ws.cell(i + 2, j + 1).value = value or ""

fn = f"declarations_{datetime.datetime.now().isoformat()[:10]}.xlsx"

print(f"\n>>> Saving as {fn}")
wb.save(fn)
